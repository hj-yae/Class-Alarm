# -*- coding: utf-8 -*-
import sys
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QFormLayout, QGridLayout, QComboBox, QLayout, QTextEdit, QDesktopWidget, QMessageBox
)
import pandas as pd 
from datetime import time
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt, QDate
from openpyxl import load_workbook, Workbook

def time_to_str(t):
    """Convert datetime.time object to string."""
    return t.strftime('%H:%M') if isinstance(t, time) else t

def load_subjects_from_excel(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='수업명')
        subjects = df['수업명'].tolist()
        return ['없음'] + subjects  # '없음'을 첫 번째 항목으로 추가
    except FileNotFoundError:
        QMessageBox.warning(None, "경고", f"파일을 찾을 수 없습니다: {file_path}")
    except ValueError:
        QMessageBox.warning(None, "경고", f"'수업명' 시트를 찾을 수 없습니다: {file_path}")
    except KeyError:
        QMessageBox.warning(None, "경고", f"'수업명' 열을 찾을 수 없습니다: {file_path}")
    except Exception as e:
        QMessageBox.critical(None, "오류", f"수업명을 불러오는 중 오류 발생: {str(e)}")
    
    # 오류 발생 시 기본 과목 목록 반환
    return ["없음", "수학", "과학", "영어", "미술", "체육", "음악", "역사"]


class ClassSettings(QWidget):
    def __init__(self, parent_setting=None, load_only=False, current_weekday_name=None, current_time=None):
        super().__init__()
        self.parent_setting = parent_setting
        self.period_editor = None
        self.timetable_editor = None
        self.notice_editor = None
        self.load_only = load_only
        self.current_weekday_name = current_weekday_name
        self.current_time = current_time
        self.period_times = {}
        self.weekly_timetable = {}
        self.notices = {}
        self.teacher_message = ""
        self.excel_file_path = self.parent_setting['excel_file_path']
        self.load_data_from_excel(self.excel_file_path)  # Load initial settings from Excel
        if not load_only:
            self.initUI()

    def initUI(self):
        self.layout = QVBoxLayout(self)
        self.layout.setSpacing(15)
        self.layout.setContentsMargins(20, 20, 20, 20)
        self.setStyleSheet("background-color: #ffffff;")

        # Period Settings
        period_button = QPushButton('수업시간 편집')
        period_button.setFont(QFont("Helvetica", 14))
        period_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 10px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        period_button.clicked.connect(self.openPeriodEditor)
        self.layout.addWidget(period_button)

        # Timetable Settings
        timetable_button = QPushButton('주간시간표 편집')
        timetable_button.setFont(QFont("Helvetica", 14))
        timetable_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                padding: 10px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        timetable_button.clicked.connect(self.openTimetableEditor)
        self.layout.addWidget(timetable_button)

        # Notice Settings
        notice_button = QPushButton('공지사항 편집')
        notice_button.setFont(QFont("Helvetica", 14))
        notice_button.setStyleSheet("""
            QPushButton {
                background-color: #ff9800;
                color: white;
                padding: 10px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #e68a00;
            }
        """)
        notice_button.clicked.connect(self.openNoticeditor)
        self.layout.addWidget(notice_button)

        # Save Button
        save_button = QPushButton('변경사항 저장')
        save_button.setFont(QFont("Helvetica", 14))
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        save_button.clicked.connect(self.save_changes)
        self.layout.addWidget(save_button)

        self.setWindowTitle("수업 설정")
        # 창 크기 고정
        self.setFixedSize(400, 250)  # 너비 400, 높이 600으로 고정
    
        # 창을 화면 중앙에 위치시키기
        self.center()
        self.show()

    def closeEvent(self, event):
        # 모든 자식 창 닫기
        if hasattr(self, 'timetable_editor') and self.timetable_editor:
            self.timetable_editor.close()
        if hasattr(self, 'notice_editor') and self.notice_editor:
            self.notice_editor.close()
        if hasattr(self, 'period_editor') and self.period_editor:
            self.period_editor.close()
        
        # 추가적인 창이 있다면 여기에 닫는 코드를 추가하세요

        event.accept()

    def load_notice(self):
        return self.notices, self.teacher_message

    def update_notice_data(self, notices, teacher_message):
        self.notices = notices
        self.teacher_message = teacher_message

    def center(self):
        # 창을 화면의 중앙에 위치시키는 메서드
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def save_changes(self):
        self.save_period_to_excel()
        self.save_class_to_excel()
        # Add any additional save logic if needed
        self.close()

    def load_data_from_excel(self, file_path):
        self.period_times = pd.read_excel(file_path, sheet_name='수업시간')
        self.period_times = {
            row['교시']: (time_to_str(row['시작시간']), time_to_str(row['종료시간']))
            for index, row in self.period_times.iterrows()
        }
        
        self.weekly_timetable = pd.read_excel(file_path, sheet_name='주간수업시간표')
        self.weekly_timetable= self.weekly_timetable.T
        self.weekly_timetable.columns = self.weekly_timetable.iloc[0]
        self.weekly_timetable = self.weekly_timetable[1:]
        self.weekly_timetable = {day: row.dropna().to_dict() for day, row in self.weekly_timetable.iterrows()}

    def save_period_to_excel(self):
        try:
            wb = load_workbook(self.excel_file_path)
        except FileNotFoundError:
            wb = Workbook()
        
        # '수업시간' 시트 선택 (없으면 새로 생성)
        if '수업시간' in wb.sheetnames:
            sheet = wb['수업시간']
            sheet.delete_rows(1, sheet.max_row)  # 기존 내용 삭제
        else:
            sheet = wb.create_sheet('수업시간')

        # 헤더 추가
        sheet.append(['교시', '시작시간', '종료시간'])

        # 데이터 추가
        for period, times in self.period_times.items():
            sheet.append([period, times[0], times[1]])

        # 파일 저장
        wb.save(self.excel_file_path)
        print(f"Period times saved to {self.excel_file_path}")
            
    def save_class_to_excel(self):
        wb = load_workbook(self.excel_file_path)
        
        # '주간수업시간표' 시트 선택 (없으면 새로 생성)
        if '주간수업시간표' in wb.sheetnames:
            sheet = wb['주간수업시간표']
            sheet.delete_rows(1, sheet.max_row)  # 기존 내용 삭제
        else:
            sheet = wb.create_sheet('주간수업시간표')

        # 요일 헤더 추가
        sheet.append(['교시'] + list(self.weekly_timetable.keys()))

        # 데이터 추가
        max_periods = max(len(day_schedule) for day_schedule in self.weekly_timetable.values())
        for period in range(1, max_periods + 1):
            row = [f"{period}교시"]
            for day in self.weekly_timetable.keys():
                subject = self.weekly_timetable[day].get(f"{period}교시", "")
                row.append(subject)
            sheet.append(row)

        # 파일 저장
        wb.save(self.excel_file_path)
        print(f"Weekly timetable saved to {self.excel_file_path}")

    def openPeriodEditor(self):
        if self.period_editor is None or not self.period_editor.isVisible():
            self.period_editor = PeriodEditor(parent_settings=self)
            self.period_editor.show()
        else:
            self.period_editor.activateWindow()
            self.period_editor.raise_()

    def openTimetableEditor(self):
        if self.timetable_editor is None or not self.timetable_editor.isVisible():
            self.timetable_editor = TimetableEditor(parent_settings=self)
            self.timetable_editor.show()
        else:
            self.timetable_editor.activateWindow()
            self.timetable_editor.raise_()

    def openNoticeditor(self):
        if self.notice_editor is None or not self.notice_editor.isVisible():
            self.notice_editor = NoticeEditor(parent_settings=self, current_weekday_name=self.current_weekday_name)
            self.notice_editor.show()
        else:
            self.notice_editor.activateWindow()
            self.notice_editor.raise_()
        
    def load_class_schedule(self):
        return self.period_times, self.weekly_timetable

class PeriodEditor(QWidget):
    def __init__(self, parent_settings):
        super().__init__()
        self.parent_settings = parent_settings
        self.initUI()

    def initUI(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                font-family: 'Helvetica';
            }
            QLabel {
                color: #333333;
            }
            QLineEdit {
                background-color: white;
                border: 1px solid #cccccc;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        # Header
        header = QLabel("수업시간 설정")
        header.setFont(QFont("Helvetica", 20, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("margin-bottom: 10px;")
        main_layout.addWidget(header)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)
        
        self.entries = {}
        
        for period, times in self.parent_settings.period_times.items():
            period_label = QLabel(period)
            period_label.setFont(QFont("Helvetica", 14, QFont.Bold))
            period_label.setAlignment(Qt.AlignCenter)
            period_label.setStyleSheet("background-color: #FF9800; color: white; padding: 5px; border-radius: 4px;")
            period_label.setFixedWidth(60)
            
            start_entry = QLineEdit(times[0])
            start_entry.setFont(QFont("Helvetica", 14))
            start_entry.setFixedWidth(100)  # 시작 시간 입력 필드의 폭을 80으로 설정
            
            end_entry = QLineEdit(times[1])
            end_entry.setFont(QFont("Helvetica", 14))
            end_entry.setFixedWidth(100)  # 종료 시간 입력 필드의 폭을 80으로 설정
            
            
            # Layout for each period
            period_layout = QHBoxLayout()
            period_layout.addWidget(QLabel("시작시간:"))
            period_layout.addWidget(start_entry)
            period_layout.addWidget(QLabel("종료시간:"))
            period_layout.addWidget(end_entry)
            
            form_layout.addRow(period_label, period_layout)
            self.entries[period] = (start_entry, end_entry)

        main_layout.addLayout(form_layout)

        # Save Button
        save_button = QPushButton('수업시간 저장')
        save_button.setFont(QFont("Helvetica", 14))
        save_button.clicked.connect(self.savePeriodTimes)
        main_layout.addWidget(save_button)

        self.setWindowTitle("수업시간 설정")
        self.setGeometry(300, 300, 500, 400)
        self.adjustSize()

    def savePeriodTimes(self):
        for period, (start_entry, end_entry) in self.entries.items():
            self.parent_settings.period_times[period] = (start_entry.text(), end_entry.text())
        self.parent_settings.save_period_to_excel()
        QMessageBox.information(self, "저장 완료", "수업 시간이 성공적으로 저장되었습니다.")
        self.close()

class TimetableEditor(QWidget):
    def __init__(self, parent_settings):
        super().__init__()
        self.parent_settings = parent_settings
        self.excel_file_path = self.parent_settings.excel_file_path
        self.subjects = load_subjects_from_excel(self.excel_file_path)
        self.days = ["월요일", "화요일", "수요일", "목요일", "금요일"]
        self.periods = ["1교시", "2교시", "3교시", "4교시", "5교시", "6교시", "7교시", "8교시"]
        self.initUI()

    def initUI(self):
        layout = QGridLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)
        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                font-family: 'Helvetica';
            }
            QLabel {
                color: #333333;
            }
            QComboBox {
                background-color: white;
                border: 1px solid #cccccc;
                border-radius: 4px;
                padding: 4px;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox::down-arrow {
                image: url(down_arrow.png);
                width: 12px;
                height: 12px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        # Header
        header = QLabel("주간시간표")
        header.setFont(QFont("Helvetica", 20, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("margin-bottom: 10px;")
        layout.addWidget(header, 0, 0, 1, len(self.days)+1)

        # Set headers for days
        for i, day in enumerate(self.days):
            day_label = QLabel(day)
            day_label.setFont(QFont("Helvetica", 12, QFont.Bold))
            day_label.setAlignment(Qt.AlignCenter)
            day_label.setStyleSheet("background-color: #2196F3; color: white; padding: 5px; border-radius: 4px;")
            day_label.setFixedWidth(80)
            layout.addWidget(day_label, 2, i+1)

        # Set headers for periods and add comboboxes
        self.comboboxes = {}
        for j, period in enumerate(self.periods):
            period_label = QLabel(period)
            period_label.setFont(QFont("Helvetica", 12, QFont.Bold))
            period_label.setAlignment(Qt.AlignCenter)
            period_label.setStyleSheet("background-color: #FF9800; color: white; padding: 5px; border-radius: 4px;")
            period_label.setFixedWidth(60)
            layout.addWidget(period_label, j+3, 0)

            for i, day in enumerate(self.days):
                combo = QComboBox()
                combo.addItems(self.subjects)
                combo.setFixedSize(80, 30)
                
                key = (day, period)
                try:
                    current_subject = self.parent_settings.weekly_timetable[day][period]
                    if current_subject in self.subjects:
                        current_index = self.subjects.index(current_subject)
                        combo.setCurrentIndex(current_index)
                    else:
                        combo.addItem(current_subject)
                        combo.setCurrentIndex(len(self.subjects))
                except KeyError:
                    combo.setCurrentIndex(0)
                    
                layout.addWidget(combo, j+3, i+1)
                self.comboboxes[key] = combo

        # 버튼 위에 추가 여백
        button_spacer = QLabel()
        button_spacer.setFixedHeight(10)
        layout.addWidget(button_spacer, len(self.periods)+3, 0, 1, len(self.days)+1)

        # Save Button
        save_button = QPushButton('주간시간표 저장')
        save_button.setFont(QFont("Helvetica", 12))
        layout.addWidget(save_button, len(self.periods)+4, 0, 1, len(self.days)+1)
        save_button.clicked.connect(self.saveTimetable)

        self.setWindowTitle("주간시간표 설정")
        
        # 창 크기를 내용에 맞게 조정
        self.adjustSize()
        
        # 창의 폭을 1.2배로 늘리고 높이는 그대로 유지
        current_width = self.width()
        new_width = int(current_width * 1.2)
        self.setFixedSize(new_width, self.height())

    def saveTimetable(self):
        for key, combo in self.comboboxes.items():
            day, period = key
            self.parent_settings.weekly_timetable[day][period] = combo.currentText()
        
        self.parent_settings.save_class_to_excel()
        QMessageBox.information(self, "저장 완료", "주간 시간표가 성공적으로 저장되었습니다.")
        self.close()

class NoticeEditor(QWidget):
    def __init__(self, parent_settings=None, current_weekday_name=None, current_time=None):
        super().__init__()
        self.parent_settings = parent_settings
        self.current_weekday_name = current_weekday_name
        self.current_time = current_time
        self.notice_entries = {}
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        self.setStyleSheet("""
            QWidget {
                background-color: #f0f0f0;
                font-family: 'Helvetica';
            }
            QLabel {
                color: #333333;
            }
            QLineEdit, QTextEdit {
                background-color: white;
                border: 1px solid #cccccc;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        # Header
        header = QLabel("공지사항 편집")
        header.setFont(QFont("Helvetica", 20, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)
        
        # 오늘의 요일 가져오기
        current_day = self.current_weekday_name
        
        if hasattr(self.parent_settings, 'weekly_timetable') and current_day in self.parent_settings.weekly_timetable:
            daily_schedule = self.parent_settings.weekly_timetable[current_day]
            for period, subject in daily_schedule.items():
                if subject == "없음":
                    continue
                notice_text = QTextEdit()
                notice_text.setPlaceholderText(f"{subject} 관련 공지사항 입력")
                notice_text.setFont(QFont("Helvetica", 12))
                
                # 기존 공지사항 불러오기
                if hasattr(self.parent_settings, 'notices') and period in self.parent_settings.notices:
                    notice_text.setText(self.parent_settings.notices[period])
                
                form_layout.addRow(QLabel(f"{period} - {subject} 공지:"), notice_text)
                self.notice_entries[period] = notice_text
        else:
            # 요일이 없을 경우 사용자에게 알림
            no_schedule_label = QLabel(f"{current_day}의 시간표가 설정되어 있지 않습니다.")
            no_schedule_label.setFont(QFont("Helvetica", 14))
            form_layout.addRow(no_schedule_label)
        
        layout.addLayout(form_layout)

        # Save Button
        save_button = QPushButton('모두 저장')
        save_button.setFont(QFont("Helvetica", 14))
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                padding: 10px;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        save_button.clicked.connect(self.saveAll)
        layout.addWidget(save_button)

        self.setWindowTitle("공지사항 관리")
        self.setGeometry(500, 200, 600, 700)
        self.show()

    def saveAll(self):
        current_day = self.current_weekday_name
        if hasattr(self.parent_settings, 'weekly_timetable') and current_day in self.parent_settings.weekly_timetable:
            self.parent_settings.notices = {period: notice_widget.toPlainText() for period, notice_widget in self.notice_entries.items()}
        print("Notices saved:", self.parent_settings.notices)
        print("Teacher's message saved:", self.parent_settings.teacher_message)
        self.parent_settings.update_notice_data(self.parent_settings.notices, self.parent_settings.teacher_message)
        self.close()